import openpyxl

def readexcel(file):
    wb = openpyxl.load_workbook(file)
    sheet = wb.active

    nrows = sheet.max_row
    ncol = sheet.max_column

    testdata = []
    for i in list(range(1, nrows)):
        for j in list(range(1, ncol)):
            tmplist = []
            tmplist.append(sheet.cell(i, j).value)
            testdata.append(tmplist)
    print(testdata)

if __name__ == "__main__":
    file = 'C:\\Users\\81288\\Desktop\\excelFile.xlsx'
    list = readexcel(file)